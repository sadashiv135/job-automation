from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOG_PATH = Path(__file__).parent / "jobs.xlsx"

COLUMNS = [
    "Date",
    "Job Title",
    "Company",
    "Location",
    "Experience Level",
    "Visa Filter",           # NEW — "OK" or "Skipped - Citizenship Required"
    "Salary",
    "Match Score (%)",
    "Job URL",
    "Resume File",
    "Cover Letter File",
    "Application Status",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT   = Font(color="0563C1", underline="single")

COL_SCORE   = COLUMNS.index("Match Score (%)") + 1    # 8
COL_URL     = COLUMNS.index("Job URL") + 1             # 9
COL_VISA    = COLUMNS.index("Visa Filter") + 1         # 6
COL_TITLE   = COLUMNS.index("Job Title") + 1           # 2
COL_COMPANY = COLUMNS.index("Company") + 1             # 3
COL_STATUS  = COLUMNS.index("Application Status") + 1  # 12

COL_WIDTHS  = [12, 30, 25, 20, 14, 28, 22, 14, 14, 40, 40, 18]

_AUTO_STATUSES = {"To Apply", "Skipped (low match)", "Skipped (citizenship)"}


def _get_or_create_workbook() -> tuple:
    """Load existing workbook or create a fresh one. Never deletes the file.
    Returns (workbook, worksheet, schema_ok).
    schema_ok=False disables upsert matching to protect old rows from column corruption."""
    if LOG_PATH.exists():
        wb = openpyxl.load_workbook(LOG_PATH)
        ws = wb.active
        existing = [ws.cell(row=1, column=i + 1).value for i in range(ws.max_column)]
        schema_ok = (existing == COLUMNS)
        if not schema_ok:
            print(
                "[logger] NOTE: jobs.xlsx has a different column schema. "
                "Old rows are preserved; new rows will be appended with updated columns. "
                "Delete jobs.xlsx for a fully clean layout."
            )
        return wb, ws, schema_ok

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    _write_header(ws)
    return wb, ws, True


def _write_header(ws) -> None:
    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _build_row_index(ws) -> dict:
    """Map normalised URL / (title, company) → row number for O(1) upsert lookups."""
    index: dict[object, int] = {}
    for row_num in range(2, ws.max_row + 1):
        href = getattr(ws.cell(row=row_num, column=COL_URL).hyperlink, "target", None) or ""
        if href:
            index[href.lower()] = row_num
        title   = ws.cell(row=row_num, column=COL_TITLE).value or ""
        company = ws.cell(row=row_num, column=COL_COMPANY).value or ""
        if title and company:
            index[(title.lower(), company.lower())] = row_num
    return index


def _find_existing_row(ws, job: dict, url: str) -> int | None:
    index = _build_row_index(ws)
    if url and url != "N/A":
        row = index.get(url.lower())
        if row:
            return row
    key = (job.get("title", "").lower(), job.get("company", "").lower())
    return index.get(key)


def _write_row(ws, row_num: int, job: dict, score: int,
               resume_path: Path, cover_letter_path: Path,
               url: str, status: str) -> None:
    """Write all fields for a row (insert or update)."""
    visa_filter = job.get("visa_filter", "OK")

    row_data = [
        datetime.now().strftime("%Y-%m-%d"),
        job.get("title", "N/A"),
        job.get("company", "N/A"),
        job.get("location", "N/A"),
        job.get("experience_level", "Entry Level"),
        visa_filter,
        job.get("salary", "Not disclosed"),
        score,
        None,   # Job URL — written below as hyperlink
        str(resume_path) if resume_path else "N/A",
        str(cover_letter_path) if cover_letter_path else "N/A",
        status,
    ]
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_idx, value=value)

    # Clickable hyperlink for Job URL
    url_cell = ws.cell(row=row_num, column=COL_URL)
    if url and url != "N/A":
        url_cell.value     = "Apply Here"
        url_cell.hyperlink = url
        url_cell.font      = LINK_FONT
        url_cell.alignment = Alignment(horizontal="center")
    else:
        url_cell.value     = "N/A"
        url_cell.hyperlink = None

    # Colour-coded match score
    score_cell = ws.cell(row=row_num, column=COL_SCORE)
    if score >= 80:
        score_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        score_cell.font = Font(color="276221", bold=True)
    elif score >= 60:
        score_cell.fill = PatternFill("solid", fgColor="FFEB9C")
        score_cell.font = Font(color="9C5700", bold=True)
    else:
        score_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        score_cell.font = Font(color="9C0006", bold=True)

    # Colour-coded visa filter
    visa_cell = ws.cell(row=row_num, column=COL_VISA)
    if visa_filter == "OK":
        visa_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        visa_cell.font = Font(color="276221", bold=True)
    else:
        visa_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        visa_cell.font = Font(color="9C0006", bold=True)


def log_job(
    job: dict,
    score: int,
    resume_path: Path,
    cover_letter_path: Path,
    status: str = "To Apply",
) -> None:
    """Upsert a job entry in jobs.xlsx. Updates existing rows, appends new ones."""
    wb, ws, schema_ok = _get_or_create_workbook()
    url = job.get("url", "N/A") or "N/A"

    # Only attempt upsert when the schema matches — prevents column corruption on old files
    existing_row = _find_existing_row(ws, job, url) if schema_ok else None

    if existing_row:
        current_status = ws.cell(row=existing_row, column=COL_STATUS).value or ""
        final_status   = status if current_status in _AUTO_STATUSES else current_status
        _write_row(ws, existing_row, job, score, resume_path, cover_letter_path, url, final_status)
        print(f"[logger] Updated row {existing_row} — {job.get('title')} @ {job.get('company')}")
    else:
        new_row = ws.max_row + 1
        _write_row(ws, new_row, job, score, resume_path, cover_letter_path, url, status)
        print(f"[logger] Appended row {new_row}  — {job.get('title')} @ {job.get('company')}")

    wb.save(LOG_PATH)
